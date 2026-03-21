"""
QuickBooks Desktop Automated Data Extractor
============================================
Connects to QuickBooks Desktop via COM SDK (QBXMLRP2) and extracts
ALL data needed for migration — no manual exports required.

Just have QuickBooks Desktop open with your company file, then run this.

Extracts:
  1. Chart of Accounts (full hierarchy with types)
  2. Customer List (with sub-customers/jobs)
  3. Vendor List (with 1099 status)
  4. Item List (service, inventory, non-inventory)
  5. Employee List
  6. Trial Balance (via GeneralSummary report)
  7. Open AR Invoices (via InvoiceQuery)
  8. Open AP Bills (via BillQuery)

Output: Excel files in ./qbd_exports/ ready for the migration agent.

Requirements:
  pip install pywin32 openpyxl

Usage:
  # Make sure QuickBooks Desktop is running with company file open
  python qbd_auto_extract.py
"""

import sys
import os
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path
import logging

# ============================================================================
# CONFIG
# ============================================================================
OUTPUT_DIR = Path("./qbd_exports")
QB_APP_NAME = "Acumatica Migration Agent"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(OUTPUT_DIR / "extract.log", mode="w"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("qbd_extract")


# ============================================================================
# QBXML REQUEST BUILDER
# ============================================================================

def wrap_qbxml(request_body, version="16.0"):
    """Wrap a request body in the QBXML envelope."""
    return f'''<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{version}"?>
<QBXML>
  <QBXMLMsgsRq onError="continueOnError">
    {request_body}
  </QBXMLMsgsRq>
</QBXML>'''


# ============================================================================
# QB CONNECTION
# ============================================================================

class QBConnection:
    """Manage connection to QuickBooks Desktop via COM."""

    def __init__(self):
        self.rp = None
        self.ticket = None

    def connect(self):
        try:
            import win32com.client
        except ImportError:
            log.error("pywin32 is not installed. Run: pip install pywin32")
            log.error("This script must run on Windows with QuickBooks Desktop installed.")
            sys.exit(1)

        log.info("Connecting to QuickBooks Desktop...")
        try:
            self.rp = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
            self.rp.OpenConnection2("", QB_APP_NAME, 1)  # 1 = localQBD
            # BeginSession: "" = current open company, 2 = doNotCare
            self.ticket = self.rp.BeginSession("", 2)
            log.info("Connected to QuickBooks Desktop successfully")
            return True
        except Exception as e:
            error_msg = str(e)
            if "80040408" in error_msg:
                log.error("QuickBooks is not running. Please open QuickBooks Desktop")
                log.error("with your company file before running this script.")
            elif "80040422" in error_msg:
                log.error("QuickBooks denied access. When the authorization dialog")
                log.error("appears in QuickBooks, click 'Yes, always allow access'.")
            elif "80040416" in error_msg:
                log.error("Another application is connected to QuickBooks in single-user mode.")
                log.error("Close other integrations or switch to multi-user mode.")
            else:
                log.error(f"Failed to connect to QuickBooks: {e}")
                log.error("Make sure QuickBooks Desktop is running with a company file open.")
            return False

    def send_request(self, qbxml):
        """Send QBXML request and return parsed XML response."""
        raw = self.rp.ProcessRequest(self.ticket, qbxml)
        return ET.fromstring(raw)

    def disconnect(self):
        if self.rp:
            try:
                if self.ticket:
                    self.rp.EndSession(self.ticket)
                self.rp.CloseConnection()
                log.info("Disconnected from QuickBooks Desktop")
            except Exception:
                pass


# ============================================================================
# XML HELPERS
# ============================================================================

def get_text(element, tag, default=""):
    """Safely get text from an XML element."""
    el = element.find(tag)
    if el is not None and el.text:
        return el.text.strip()
    return default


def get_nested_text(element, parent_tag, child_tag, default=""):
    """Get text from a nested element like <ParentRef><FullName>."""
    parent = element.find(parent_tag)
    if parent is not None:
        return get_text(parent, child_tag, default)
    return default


def get_address(element, addr_tag):
    """Extract address fields from a QBD address block."""
    addr = element.find(addr_tag)
    if addr is None:
        return {"Addr1": "", "Addr2": "", "City": "", "State": "", "PostalCode": "", "Country": ""}
    return {
        "Addr1": get_text(addr, "Addr1"),
        "Addr2": get_text(addr, "Addr2"),
        "City": get_text(addr, "City"),
        "State": get_text(addr, "State"),
        "PostalCode": get_text(addr, "PostalCode"),
        "Country": get_text(addr, "Country", "US"),
    }


# ============================================================================
# EXTRACTORS — one per data type
# ============================================================================

def extract_chart_of_accounts(qb):
    """Extract all GL accounts with types and hierarchy."""
    log.info("Extracting chart of accounts...")
    request = wrap_qbxml("""<AccountQueryRq>
        <ActiveStatus>All</ActiveStatus>
    </AccountQueryRq>""")

    root = qb.send_request(request)
    accounts = []
    for acct in root.iter("AccountRet"):
        accounts.append({
            "ListID": get_text(acct, "ListID"),
            "Name": get_text(acct, "Name"),
            "FullName": get_text(acct, "FullName"),
            "AccountType": get_text(acct, "AccountType"),
            "AccountNumber": get_text(acct, "AccountNumber"),
            "Description": get_text(acct, "Desc"),
            "Balance": get_text(acct, "Balance", "0"),
            "IsActive": get_text(acct, "IsActive", "true"),
            "ParentFullName": get_nested_text(acct, "ParentRef", "FullName"),
            "Sublevel": get_text(acct, "Sublevel", "0"),
            "BankNumber": get_text(acct, "BankNumber"),
        })
    log.info(f"  Extracted {len(accounts)} accounts")
    return accounts


def extract_customers(qb):
    """Extract all customers and sub-customers/jobs."""
    log.info("Extracting customers...")
    request = wrap_qbxml("""<CustomerQueryRq>
        <ActiveStatus>All</ActiveStatus>
        <OwnerID>0</OwnerID>
    </CustomerQueryRq>""")

    root = qb.send_request(request)
    customers = []
    for cust in root.iter("CustomerRet"):
        bill_addr = get_address(cust, "BillAddress")
        ship_addr = get_address(cust, "ShipAddress")
        customers.append({
            "ListID": get_text(cust, "ListID"),
            "Name": get_text(cust, "Name"),
            "FullName": get_text(cust, "FullName"),
            "IsActive": get_text(cust, "IsActive", "true"),
            "CompanyName": get_text(cust, "CompanyName"),
            "Salutation": get_text(cust, "Salutation"),
            "FirstName": get_text(cust, "FirstName"),
            "LastName": get_text(cust, "LastName"),
            "Phone": get_text(cust, "Phone"),
            "AltPhone": get_text(cust, "AltPhone"),
            "Fax": get_text(cust, "Fax"),
            "Email": get_text(cust, "Email"),
            "BillAddr1": bill_addr["Addr1"],
            "BillAddr2": bill_addr["Addr2"],
            "BillCity": bill_addr["City"],
            "BillState": bill_addr["State"],
            "BillZip": bill_addr["PostalCode"],
            "ShipAddr1": ship_addr["Addr1"],
            "ShipCity": ship_addr["City"],
            "ShipState": ship_addr["State"],
            "ShipZip": ship_addr["PostalCode"],
            "Terms": get_nested_text(cust, "TermsRef", "FullName"),
            "Balance": get_text(cust, "Balance", "0"),
            "TotalBalance": get_text(cust, "TotalBalance", "0"),
            "ParentFullName": get_nested_text(cust, "ParentRef", "FullName"),
            "JobStatus": get_text(cust, "JobStatus"),
            "JobType": get_nested_text(cust, "JobTypeRef", "FullName"),
            "JobDesc": get_text(cust, "JobDesc"),
            "JobStartDate": get_text(cust, "JobStartDate"),
            "JobProjectedEndDate": get_text(cust, "JobProjectedEndDate"),
            "JobEndDate": get_text(cust, "JobEndDate"),
            "SalesTaxCode": get_nested_text(cust, "SalesTaxCodeRef", "FullName"),
            "CustomerType": get_nested_text(cust, "CustomerTypeRef", "FullName"),
        })
    log.info(f"  Extracted {len(customers)} customers (including sub-customers/jobs)")
    return customers


def extract_vendors(qb):
    """Extract all vendors with 1099 status."""
    log.info("Extracting vendors...")
    request = wrap_qbxml("""<VendorQueryRq>
        <ActiveStatus>All</ActiveStatus>
        <OwnerID>0</OwnerID>
    </VendorQueryRq>""")

    root = qb.send_request(request)
    vendors = []
    for vend in root.iter("VendorRet"):
        addr = get_address(vend, "VendorAddress")
        vendors.append({
            "ListID": get_text(vend, "ListID"),
            "Name": get_text(vend, "Name"),
            "IsActive": get_text(vend, "IsActive", "true"),
            "CompanyName": get_text(vend, "CompanyName"),
            "Salutation": get_text(vend, "Salutation"),
            "FirstName": get_text(vend, "FirstName"),
            "LastName": get_text(vend, "LastName"),
            "Phone": get_text(vend, "Phone"),
            "AltPhone": get_text(vend, "AltPhone"),
            "Fax": get_text(vend, "Fax"),
            "Email": get_text(vend, "Email"),
            "Addr1": addr["Addr1"],
            "Addr2": addr["Addr2"],
            "City": addr["City"],
            "State": addr["State"],
            "PostalCode": addr["PostalCode"],
            "Terms": get_nested_text(vend, "TermsRef", "FullName"),
            "Balance": get_text(vend, "Balance", "0"),
            "Is1099": get_text(vend, "Is1099", "false"),
            "TaxID": get_text(vend, "VendorTaxIdent"),
            "AccountNumber": get_text(vend, "AccountNumber"),
            "VendorType": get_nested_text(vend, "VendorTypeRef", "FullName"),
        })
    log.info(f"  Extracted {len(vendors)} vendors")
    return vendors


def extract_items(qb):
    """Extract all items (service, inventory, non-inventory, other charge)."""
    log.info("Extracting items...")
    request = wrap_qbxml("""<ItemQueryRq>
        <ActiveStatus>All</ActiveStatus>
        <OwnerID>0</OwnerID>
    </ItemQueryRq>""")

    root = qb.send_request(request)
    items = []
    item_types = [
        "ItemServiceRet", "ItemInventoryRet", "ItemNonInventoryRet",
        "ItemOtherChargeRet", "ItemGroupRet", "ItemFixedAssetRet",
        "ItemDiscountRet", "ItemPaymentRet", "ItemSalesTaxRet",
        "ItemSalesTaxGroupRet", "ItemSubtotalRet",
    ]
    for item_type in item_types:
        for item in root.iter(item_type):
            clean_type = item_type.replace("Item", "").replace("Ret", "")
            items.append({
                "ListID": get_text(item, "ListID"),
                "Name": get_text(item, "Name"),
                "FullName": get_text(item, "FullName"),
                "ItemType": clean_type,
                "IsActive": get_text(item, "IsActive", "true"),
                "Description": get_text(item, "Desc") or get_text(item, "SalesDesc"),
                "Price": get_text(item, "SalesPrice") or get_text(item, "Price", "0"),
                "Cost": get_text(item, "PurchaseCost") or get_text(item, "Cost", "0"),
                "IncomeAccount": get_nested_text(item, "IncomeAccountRef", "FullName"),
                "ExpenseAccount": get_nested_text(item, "ExpenseAccountRef", "FullName"),
                "COGSAccount": get_nested_text(item, "COGSAccountRef", "FullName"),
                "AssetAccount": get_nested_text(item, "AssetAccountRef", "FullName"),
                "ParentFullName": get_nested_text(item, "ParentRef", "FullName"),
            })
    log.info(f"  Extracted {len(items)} items")
    return items


def extract_employees(qb):
    """Extract employee list."""
    log.info("Extracting employees...")
    request = wrap_qbxml("""<EmployeeQueryRq>
        <ActiveStatus>All</ActiveStatus>
        <OwnerID>0</OwnerID>
    </EmployeeQueryRq>""")

    root = qb.send_request(request)
    employees = []
    for emp in root.iter("EmployeeRet"):
        addr = get_address(emp, "EmployeeAddress")
        employees.append({
            "ListID": get_text(emp, "ListID"),
            "Name": get_text(emp, "Name"),
            "IsActive": get_text(emp, "IsActive", "true"),
            "FirstName": get_text(emp, "FirstName"),
            "MiddleName": get_text(emp, "MiddleName"),
            "LastName": get_text(emp, "LastName"),
            "SSN": get_text(emp, "SSN"),
            "Phone": get_text(emp, "Phone"),
            "Mobile": get_text(emp, "Mobile"),
            "Email": get_text(emp, "Email"),
            "Addr1": addr["Addr1"],
            "City": addr["City"],
            "State": addr["State"],
            "PostalCode": addr["PostalCode"],
            "HiredDate": get_text(emp, "HiredDate"),
            "ReleasedDate": get_text(emp, "ReleasedDate"),
            "BirthDate": get_text(emp, "BirthDate"),
            "EmployeeType": get_text(emp, "EmployeeType"),
            "Gender": get_text(emp, "Gender"),
        })
    log.info(f"  Extracted {len(employees)} employees")
    return employees


def extract_open_invoices(qb):
    """Extract all open (unpaid/partially paid) AR invoices."""
    log.info("Extracting open AR invoices...")
    request = wrap_qbxml("""<InvoiceQueryRq>
        <PaidStatus>NotPaidOnly</PaidStatus>
        <IncludeLineItems>true</IncludeLineItems>
    </InvoiceQueryRq>""")

    root = qb.send_request(request)
    invoices = []
    for inv in root.iter("InvoiceRet"):
        lines = []
        for line in inv.iter("InvoiceLineRet"):
            lines.append({
                "ItemName": get_nested_text(line, "ItemRef", "FullName"),
                "Description": get_text(line, "Desc"),
                "Quantity": get_text(line, "Quantity", "0"),
                "Rate": get_text(line, "Rate", "0"),
                "Amount": get_text(line, "Amount", "0"),
                "Account": get_nested_text(line, "AccountRef", "FullName"),
            })
        invoices.append({
            "TxnID": get_text(inv, "TxnID"),
            "RefNumber": get_text(inv, "RefNumber"),
            "CustomerFullName": get_nested_text(inv, "CustomerRef", "FullName"),
            "TxnDate": get_text(inv, "TxnDate"),
            "DueDate": get_text(inv, "DueDate"),
            "Amount": get_text(inv, "Subtotal", "0"),
            "BalanceRemaining": get_text(inv, "BalanceRemaining", "0"),
            "IsPaid": get_text(inv, "IsPaid", "false"),
            "Terms": get_nested_text(inv, "TermsRef", "FullName"),
            "PONumber": get_text(inv, "PONumber"),
            "Memo": get_text(inv, "Memo"),
            "Lines": lines,
        })
    log.info(f"  Extracted {len(invoices)} open invoices")
    return invoices


def extract_open_bills(qb):
    """Extract all open (unpaid/partially paid) AP bills."""
    log.info("Extracting open AP bills...")
    request = wrap_qbxml("""<BillQueryRq>
        <PaidStatus>NotPaidOnly</PaidStatus>
        <IncludeLineItems>true</IncludeLineItems>
    </BillQueryRq>""")

    root = qb.send_request(request)
    bills = []
    for bill in root.iter("BillRet"):
        lines = []
        for line in bill.iter("ExpenseLineRet"):
            lines.append({
                "Account": get_nested_text(line, "AccountRef", "FullName"),
                "Amount": get_text(line, "Amount", "0"),
                "Memo": get_text(line, "Memo"),
                "Customer": get_nested_text(line, "CustomerRef", "FullName"),
            })
        for line in bill.iter("ItemLineRet"):
            lines.append({
                "ItemName": get_nested_text(line, "ItemRef", "FullName"),
                "Amount": get_text(line, "Amount", "0"),
                "Quantity": get_text(line, "Quantity", "0"),
                "Cost": get_text(line, "Cost", "0"),
                "Customer": get_nested_text(line, "CustomerRef", "FullName"),
            })
        bills.append({
            "TxnID": get_text(bill, "TxnID"),
            "RefNumber": get_text(bill, "RefNumber"),
            "VendorFullName": get_nested_text(bill, "VendorRef", "FullName"),
            "TxnDate": get_text(bill, "TxnDate"),
            "DueDate": get_text(bill, "DueDate"),
            "AmountDue": get_text(bill, "AmountDue", "0"),
            "IsPaid": get_text(bill, "IsPaid", "false"),
            "Terms": get_nested_text(bill, "TermsRef", "FullName"),
            "Memo": get_text(bill, "Memo"),
            "Lines": lines,
        })
    log.info(f"  Extracted {len(bills)} open bills")
    return bills


def extract_trial_balance(qb):
    """Extract account balances via AccountQuery (Balance field)."""
    log.info("Extracting trial balance (account balances)...")
    # We already get Balance in the COA extract, but let's also get
    # a clean summary. The GeneralSummaryReport gives us proper
    # debit/credit columns.
    request = wrap_qbxml(f"""<GeneralSummaryReportQueryRq>
        <GeneralSummaryReportType>TrialBalance</GeneralSummaryReportType>
        <ReportPeriod>
            <FromReportDate>{date.today().strftime('%Y-01-01')}</FromReportDate>
            <ToReportDate>{date.today().strftime('%Y-%m-%d')}</ToReportDate>
        </ReportPeriod>
    </GeneralSummaryReportQueryRq>""")

    root = qb.send_request(request)
    rows = []
    for report in root.iter("ReportRet"):
        for data_row in report.iter("DataRow"):
            cols = list(data_row.iter("ColData"))
            if len(cols) >= 3:
                rows.append({
                    "Account": cols[0].get("value", ""),
                    "Debit": cols[1].get("value", "0"),
                    "Credit": cols[2].get("value", "0"),
                })
    log.info(f"  Extracted {len(rows)} trial balance rows")
    return rows


# ============================================================================
# EXCEL WRITER
# ============================================================================

def save_to_excel(data, filename, sheet_name="Sheet1"):
    """Save list of dicts to Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    filepath = OUTPUT_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws.cell(row=1, column=1, value="No data extracted")
        wb.save(str(filepath))
        return filepath

    # Flatten nested data (like invoice lines)
    flat_data = []
    for row in data:
        flat_row = {}
        for k, v in row.items():
            if isinstance(v, list):
                flat_row[f"{k}_Count"] = len(v)
                # Serialize lines as JSON-ish string for reference
                if v:
                    flat_row[f"{k}_Summary"] = "; ".join(
                        str({kk: vv for kk, vv in line.items() if vv})
                        for line in v[:5]
                    )
            else:
                flat_row[k] = v
        flat_data.append(flat_row)

    # Get all unique headers
    headers = []
    for row in flat_data:
        for k in row:
            if k not in headers:
                headers.append(k)

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for r, row in enumerate(flat_data, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(h, ""))

    # Auto-width columns
    for c, h in enumerate(headers, 1):
        max_len = max(len(str(h)), max((len(str(row.get(h, ""))) for row in flat_data), default=0))
        ws.column_dimensions[chr(64 + c) if c <= 26 else "A"].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(str(filepath))
    log.info(f"  Saved: {filepath}")
    return filepath


# ============================================================================
# MAIN
# ============================================================================

def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    log.info("=" * 60)
    log.info("  QuickBooks Desktop — Automated Data Extractor")
    log.info("=" * 60)
    log.info(f"  Output: {OUTPUT_DIR.resolve()}")
    log.info(f"  Time:   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)

    qb = QBConnection()
    if not qb.connect():
        log.info("\n" + "=" * 60)
        log.info("QUICKBOOKS CONNECTION FAILED")
        log.info("=" * 60)
        log.info("Checklist:")
        log.info("  1. Is QuickBooks Desktop running?")
        log.info("  2. Is a company file open?")
        log.info("  3. When QB asks to authorize this app, click 'Yes, always'")
        log.info("  4. Is QB in single-user or multi-user mode? (both work)")
        log.info("  5. Is another integration blocking the connection?")
        sys.exit(1)

    try:
        # Extract all data types
        coa      = extract_chart_of_accounts(qb)
        customers = extract_customers(qb)
        vendors   = extract_vendors(qb)
        items     = extract_items(qb)
        employees = extract_employees(qb)
        invoices  = extract_open_invoices(qb)
        bills     = extract_open_bills(qb)
        trial_bal = extract_trial_balance(qb)

        # Save to Excel files
        log.info("\nSaving to Excel files...")
        save_to_excel(coa,       "QBD_ChartOfAccounts.xlsx", "Accounts")
        save_to_excel(customers, "QBD_Customers.xlsx",       "Customers")
        save_to_excel(vendors,   "QBD_Vendors.xlsx",         "Vendors")
        save_to_excel(items,     "QBD_Items.xlsx",           "Items")
        save_to_excel(employees, "QBD_Employees.xlsx",       "Employees")
        save_to_excel(invoices,  "QBD_OpenInvoices.xlsx",    "Open Invoices")
        save_to_excel(bills,     "QBD_OpenBills.xlsx",       "Open Bills")
        save_to_excel(trial_bal, "QBD_TrialBalance.xlsx",    "Trial Balance")

    finally:
        qb.disconnect()

    # Summary
    log.info("\n" + "=" * 60)
    log.info("  EXTRACTION COMPLETE")
    log.info("=" * 60)
    log.info(f"  Chart of Accounts:  {len(coa)} accounts")
    log.info(f"  Customers:          {len(customers)} (incl. sub-customers)")
    log.info(f"  Vendors:            {len(vendors)}")
    log.info(f"  Items:              {len(items)}")
    log.info(f"  Employees:          {len(employees)}")
    log.info(f"  Open Invoices (AR): {len(invoices)}")
    log.info(f"  Open Bills (AP):    {len(bills)}")
    log.info(f"  Trial Balance:      {len(trial_bal)} rows")
    log.info(f"\n  Files saved to: {OUTPUT_DIR.resolve()}")
    log.info("=" * 60)
    log.info("\nNext step: run the migration agent:")
    log.info(f"  python migration_agent.py --from-files")


if __name__ == "__main__":
    main()
